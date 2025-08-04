from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                            QHBoxLayout, QLabel, QPushButton, QFileDialog, QComboBox,
                            QTableView, QHeaderView, QMessageBox, QInputDialog, QListWidget,
                            QDialog, QCheckBox, QListWidgetItem, QTabWidget, QScrollArea, 
                            QFormLayout, QSpinBox, QDialogButtonBox, QSlider, QToolTip)
from PyQt5.QtCore import Qt, QTimer, QThread, pyqtSignal, QEvent
from PyQt5.QtGui import QImage, QPixmap, QIcon, QStandardItem, QFont, QColor, QPalette, QBrush, QPen, QStandardItemModel
import sys
import cv2
from ultralytics import YOLO
import numpy as np
import os
from datetime import datetime
import pandas as pd
import openpyxl
import json  # 添加json模块用于保存线段数据
# --- Import from config file --- 
from analysis_config import (player_id_mapping, recognize_action, 
                            recognize_action_with_temporal, reset_temporal_analyzers)
from detection_config import filter_detections, ActionCooldownManager
# 导入划线工具
from 划线工具 import LineAnnotator, annotate_court_lines

# 获取应用程序的根目录
def get_app_root():
    """获取应用程序的根目录，兼容开发环境和打包后的环境"""
    if getattr(sys, 'frozen', False):
        # 如果是打包后的应用程序
        return os.path.dirname(sys.executable)
    else:
        # 如果是开发环境
        return os.path.dirname(os.path.abspath(__file__))

class VolleyballAnalyzer(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("排球视频分析系统/Volleyball Video Analysis System")
        self.setGeometry(100, 100, 1600, 1000)
        
        # 使用相对路径加载图标
        app_root = get_app_root()
        icon_path = os.path.join(app_root, 'img', 'logo.png')
        self.setWindowIcon(QIcon(icon_path))
        
        self.setStyleSheet("background-color: rgb(108, 155, 190); color: #ffffff;")

        # ID 映射关系 (类别索引 -> 球员ID) - MOVED TO analysis_config.py
        # self.player_id_mapping = {
        #     0: 1, 1: 18, 2: 17, 3: 13, 4: 8, 
        #     5: 11, 6: 5, 7: 14, 8: 6
        # }

        # 初始化模型
        self.pose_model = YOLO("./model/yolo11s-pose.pt")
        self.detect_model = YOLO("./model/best.pt")

        # 初始化视频相关变量
        self.cap = None
        self.video_source = None
        self.frame_count = 0
        self.current_video_path = None  # 添加变量保存当前视频路径
        
        self.log_file = None
        
        # 跟踪已识别的球员ID
        self.all_detected_player_ids = [] # 存储所有检测到的球员ID
        self.current_displayed_player_ids = [] # 当前表格中显示的ID
        self.max_players_to_display = 10 # 限制表格显示最多10个球员
        
        # --- 新增：存储每帧的动作信息 --- 
        # 结构: {frame_num: {player_id: action_label, ...}, ...}
        self.frame_actions = {}
        
        # --- 新增：动作冷却管理器 ---
        self.action_cooldown_manager = ActionCooldownManager()
        
        # --- 新增：存储用户绘制的线段 ---
        self.court_lines = []
        
        # --- 新增：定义要连接的三等分点对 ---
        self.connection_pairs = [("2.1", "4.2"), ("2.2", "4.1"), ("1.2", "3.1"), ("1.1", "3.2")]
        
        # --- 新增：线段历史管理目录 ---
        self.lines_history_dir = os.path.join(get_app_root(), "lines_history")
        if not os.path.exists(self.lines_history_dir):
            os.makedirs(self.lines_history_dir)
        
        # 添加到__init__方法中
        self.player_area_actions = {}  # 结构: {frame_num: {player_id: (area_num, action_label), ...}, ...}
        self.area_action_stats = {i: {} for i in range(1, 10)}  # 结构: {area_num: {action_label: count, ...}, ...}
        
        self.video_writer = None  # 新增：视频写入器
        self.output_video_path = None  # 新增：输出视频路径
        
        # --- 新增：时序分析相关变量 ---
        self.use_temporal_analysis = False  # 默认关闭时序分析
        
        # --- 新增：时间轴和视频控制变量 ---
        self.video_fps = 30.0  # 视频帧率
        self.video_duration = 0  # 视频总时长（秒）
        self.seek_frame = -1  # 跳转目标帧（-1表示不跳转）
        
        self.setup_ui()
        
    def setup_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)
        main_layout.setContentsMargins(0, 0, 0, 0) # Remove padding for the main layout
        main_layout.setSpacing(0) # Remove spacing

        # --- 标题 --- 
        title_label = QLabel("排球视频分析系统/Volleyball Video Analysis System")
        title_label.setFont(QFont("Arial", 18, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        # Slightly adjusted title style
        title_label.setStyleSheet("color: #ffffff; background-color: rgba(255, 255, 255, 100); padding: 10px 0;") 
        main_layout.addWidget(title_label)
        
        # --- 创建标签页控件 --- 
        self.tab_widget = QTabWidget()
        self.tab_widget.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #8a9ba8;
                background-color: rgb(108, 155, 190);
            }
            QTabBar::tab {
                background-color: rgba(255, 255, 255, 180);
                color: #2c3e50;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 4px;
                border-top-right-radius: 4px;
            }
            QTabBar::tab:selected {
                background-color: rgb(108, 155, 190);
                color: #ffffff;
                border-bottom: 2px solid #ffffff;
            }
            QTabBar::tab:hover {
                background-color: rgba(255, 255, 255, 220);
            }
        """)
        
        # 创建基础功能标签页
        self.setup_basic_tab()
        
        # 创建高级功能标签页
        self.setup_advanced_tab()
        
        # 创建扩展功能标签页
        self.setup_extended_tab()
        
        main_layout.addWidget(self.tab_widget)
        
    def setup_basic_tab(self):
        """设置基础功能标签页"""
        basic_widget = QWidget()
        basic_layout = QVBoxLayout(basic_widget)
        basic_layout.setContentsMargins(0, 0, 0, 0)
        basic_layout.setSpacing(0)

        # --- 按钮区域 --- 
        button_layout_widget = QWidget() # Use a widget for background color
        button_layout_widget.setStyleSheet("background-color: rgba(255, 255, 255, 120); border-bottom: 1px solid #8a9ba8;")
        button_layout = QHBoxLayout(button_layout_widget)
        button_layout.setSpacing(15)
        button_layout.setContentsMargins(20, 10, 20, 10) 

        self.source_combo = QComboBox()
        self.source_combo.addItems(["视频文件/Video File", "相机/Camera"])
        self.source_combo.currentIndexChanged.connect(self.source_changed)
        
        self.load_button = QPushButton("加载视频/Load Video")
        self.start_button = QPushButton("开始分析/Start Analysis")
        self.pause_button = QPushButton("暂停/Pause")
        self.pause_button.setEnabled(False)
        self.stop_button = QPushButton("停止/Stop")
        self.export_button = QPushButton("输出报表/Export Report")
        self.export_button.clicked.connect(self.export_report)
        
        # 添加线段管理按钮
        self.manage_lines_button = QPushButton("管理线段/Manage Lines")
        self.manage_lines_button.clicked.connect(self.manage_court_lines)

        # Button/Combo Styles (keep as before)
        button_style = """
            QPushButton {
                background-color: rgba(255, 255, 255, 200); /* Match main background */
                color: #2c3e50;
                border: 1px solid #2c3e50;
                padding: 8px 15px;
                border-radius: 4px;
                font-size: 13px; /* Slightly smaller font */
            }
            QPushButton:hover {
                background-color: rgba(255, 255, 255, 255);
            }
            QPushButton:pressed {
                background-color: rgba(255, 255, 255, 150);
            }
        """
        combo_style = """
            QComboBox {
                background-color: rgba(255, 255, 255, 200);
                color: #2c3e50;
                border: 1px solid #2c3e50;
                padding: 8px 15px;
                border-radius: 4px;
                font-size: 13px;
            }
            QComboBox::drop-down {
                border: none;
                width: 20px;
            }
            QComboBox::down-arrow {
                 image: url(img/down_arrow_mint.png); /* 使用相对路径 */
                 width: 12px;
                 height: 12px;
            }
             QComboBox QAbstractItemView {
                background-color: rgba(255, 255, 255, 240);
                border: 1px solid #2c3e50;
                color: #2c3e50;
                selection-background-color: rgba(255, 255, 255, 180); /* Lighter selection */
                padding: 5px;
                outline: 0px; /* Remove focus outline */
            }
        """

        self.source_combo.setStyleSheet(combo_style)
        for button in [self.load_button, self.start_button, self.pause_button, self.stop_button, self.export_button, self.manage_lines_button]:
            button.setStyleSheet(button_style)
            button.setCursor(Qt.PointingHandCursor)

        button_layout.addWidget(self.source_combo)
        button_layout.addWidget(self.load_button)
        button_layout.addWidget(self.start_button)
        button_layout.addWidget(self.pause_button)
        button_layout.addWidget(self.stop_button)
        button_layout.addWidget(self.export_button)
        button_layout.addWidget(self.manage_lines_button)
        button_layout.addStretch()
        
        basic_layout.addWidget(button_layout_widget)
        
        # --- 内容区域 --- 
        content_layout = QHBoxLayout()
        content_layout.setSpacing(20) # Spacing between video and table area
        content_layout.setContentsMargins(20, 20, 20, 20) # Margins for the content area
        basic_layout.addLayout(content_layout, 1) # Give content layout stretch factor

        # --- 视频显示区域 (左侧) ---
        video_display_layout = QVBoxLayout()
        video_display_layout.setSpacing(15) # Spacing between original and combined video

        # 创建标签容器 (垂直布局)
        original_container = QVBoxLayout()
        combined_container = QVBoxLayout()
        
        # 创建说明标签
        original_text = QLabel("原始视频/Original Video")
        combined_text = QLabel("综合分析/Combined Analysis")

        # 标签样式 (slightly adjusted)
        label_style = """
            QLabel {
                color: #2c3e50;
                background-color: rgba(255, 255, 255, 180);
                padding: 6px 10px;
                border-top-left-radius: 4px; 
                border-top-right-radius: 4px;
                font-size: 14px;
                border: 1px solid #8a9ba8; 
                border-bottom: none; 
                font-weight: bold;
            }
        """
        for label in [original_text, combined_text]:
            label.setStyleSheet(label_style)
            # label.setAlignment(Qt.AlignCenter) # Keep left-aligned
        
        # 创建视频显示标签
        self.original_label = QLabel()
        self.combined_label = QLabel()

        # 视频标签样式 (kept same)
        video_label_style = """
            QLabel {
                border: 1px solid #8a9ba8;
                background-color: #000000; /* Black background for video */
                 border-bottom-left-radius: 4px; 
                 border-bottom-right-radius: 4px;
                 border-top-right-radius: 4px;
            }
        """
        for label in [self.original_label, self.combined_label]:
            label.setAlignment(Qt.AlignCenter)
            # label.setFixedSize(fixed_width, fixed_height) # Remove fixed size, let it scale
            label.setMinimumSize(400, 300) # Set minimum size
            label.setStyleSheet(video_label_style)

        # 组装视频和标签 (垂直布局)
        original_container.addWidget(original_text)
        original_container.addWidget(self.original_label, 1) # Give label stretch factor
        original_container.setSpacing(0)
        
        combined_container.addWidget(combined_text)
        combined_container.addWidget(self.combined_label, 1) # Give label stretch factor
        combined_container.setSpacing(0)

        # 添加到左侧视频布局
        video_display_layout.addLayout(original_container)
        video_display_layout.addLayout(combined_container)
        
        # --- 添加时间轴组件 ---
        self.timeline_widget = EventTimelineWidget()
        video_display_layout.addWidget(self.timeline_widget)
        
        # --- 右侧表格区域 ---
        right_panel_layout = QVBoxLayout()
        right_panel_layout.setSpacing(10)

        # 表格标题
        table_title = QLabel("A队伍 vs B队伍") 
        table_title.setFont(QFont("Arial", 16, QFont.Bold))
        table_title.setAlignment(Qt.AlignCenter)
        table_title.setStyleSheet("color: #ffffff; padding: 10px; background-color: rgba(255, 255, 255, 120); border-radius: 4px;")
        right_panel_layout.addWidget(table_title)

        # 创建表格视图和模型
        self.analysis_table = QTableView()
        self.table_model = QStandardItemModel()
        self.analysis_table.setModel(self.table_model)

        # 设置初始表头数据 (will be updated dynamically)
        self.initial_h_headers = ["动作/ID", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"] # Placeholder headers for 10 players
        action_labels = ["Serve", "Reception", "Attack", "Block", "Dig", "Set"]  # 移除了"触网"和"球出界"
        self.table_model.setColumnCount(len(self.initial_h_headers))
        self.table_model.setRowCount(len(action_labels))
        self.table_model.setHorizontalHeaderLabels(self.initial_h_headers)

        # 填充第一列动作名称和初始化其他单元格
        action_font = QFont()
        action_font.setBold(True)
        for row, action in enumerate(action_labels):
            # First column (Action Names)
            action_item = QStandardItem(action)
            action_item.setEditable(False)
            action_item.setTextAlignment(Qt.AlignCenter) # Center align action names like header
            action_item.setForeground(QColor("#64ffda")) # Mint color like header
            # action_item.setBackground(QColor("#112240")) # Match header background - CSS handles this better
            action_item.setFont(action_font) # Make action names bold
            self.table_model.setItem(row, 0, action_item)
            
            # Other columns (Player IDs 1-6)
            for col in range(1, len(self.initial_h_headers)):
                item = QStandardItem("") # Initialize empty
                item.setEditable(False) 
                item.setTextAlignment(Qt.AlignCenter)
                item.setForeground(QColor("#ccd6f6")) # Default text color for data area
                self.table_model.setItem(row, col, item)

        # 应用表格样式 (Refined)
        table_style = """
            QTableView {
                background-color: rgba(255, 255, 255, 220); 
                color: #2c3e50; 
                border: 1px solid #8a9ba8; 
                gridline-color: #8a9ba8; 
                border-radius: 4px;
                font-size: 13px;
            }
            QTableView::item {
                padding: 8px 5px; 
                border-bottom: 1px solid #8a9ba8; 
                border-right: 1px solid #8a9ba8; 
            }
            /* Style for the first column to look like a header */
            QTableView::item:vertical:first-child { /* Targeting items in the first column */
                 color: #2c3e50; 
                 background-color: rgba(255, 255, 255, 180); /* Match header background */
                 font-weight: bold;
            }
            QTableView::item:selected {
                background-color: rgba(255, 255, 255, 255); 
                color: #2c3e50; 
            }
            QHeaderView::section {
                background-color: rgba(255, 255, 255, 180); 
                color: #2c3e50; 
                padding: 8px 5px; /* Match item padding */
                border: none; 
                border-bottom: 1px solid #8a9ba8; 
                 /* border-right: 1px solid #8a9ba8; */ /* Apply only to horizontal? */
                font-weight: bold;
                font-size: 14px;
            }
             QHeaderView::section:horizontal {
                 border-right: 1px solid #8a9ba8;
                 text-align: center; /* Ensure header text is centered */
             }
             /* Remove right border for the last horizontal header section */
             QHeaderView::section:horizontal:last {
                 border-right: none;
             }
             /* Vertical header is hidden, but keep style just in case */
             QHeaderView::section:vertical {
                /* border-right: 1px solid #8a9ba8; */ /* Hide border if header is hidden */
                border-right: none;
             }
             QTableView QTableCornerButton::section {
                 background-color: rgba(255, 255, 255, 180);
                 border: 1px solid #8a9ba8;
                 border-bottom: 1px solid #8a9ba8;
                 border-right: 1px solid #8a9ba8;
             }
        """
        self.analysis_table.setStyleSheet(table_style)

        # 调整表格行为
        self.analysis_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch) 
        self.analysis_table.verticalHeader().setSectionResizeMode(QHeaderView.Stretch) 
        self.analysis_table.setEditTriggers(QTableView.NoEditTriggers) 
        self.analysis_table.setSelectionBehavior(QTableView.SelectItems)
        self.analysis_table.setSelectionMode(QTableView.SingleSelection)
        self.analysis_table.setShowGrid(True) 
        self.analysis_table.setAlternatingRowColors(False)

        # Hide default vertical header numbers
        self.analysis_table.verticalHeader().setVisible(False)

        # Set first column width and behavior specifically
        self.analysis_table.setColumnWidth(0, 120) # Fixed width for action labels
        self.analysis_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        self.analysis_table.horizontalHeader().setDefaultAlignment(Qt.AlignCenter) # Center align all horizontal headers

        right_panel_layout.addWidget(self.analysis_table, 1)

        # 添加到主内容布局
        # Adjust stretch factors again: Video (3) vs Table (4) to give table more space
        content_layout.addLayout(video_display_layout, 3) 
        content_layout.addLayout(right_panel_layout, 4) 
        
        # 连接信号和槽
        self.load_button.clicked.connect(self.load_video)
        self.start_button.clicked.connect(self.start_analysis)
        self.pause_button.clicked.connect(self.pause_analysis)
        self.stop_button.clicked.connect(self.stop_analysis)
        
        # 初始化定时器
        self.timer = QTimer()
        self.timer.timeout.connect(self.update_frame)
        
        # 将基础功能标签页添加到标签页控件
        self.tab_widget.addTab(basic_widget, "基础分析/Basic Analysis")
        
    def setup_advanced_tab(self):
        """设置高级功能标签页"""
        advanced_widget = QWidget()
        advanced_layout = QVBoxLayout(advanced_widget)
        advanced_layout.setContentsMargins(20, 20, 20, 20)
        
        # 高级功能说明
        info_label = QLabel("高级功能/Advanced Functions")
        info_label.setFont(QFont("Arial", 16, QFont.Bold))
        info_label.setAlignment(Qt.AlignCenter)
        info_label.setStyleSheet("color: #ffffff; padding: 20px; background-color: rgba(255, 255, 255, 120); border-radius: 4px; margin-bottom: 20px;")
        advanced_layout.addWidget(info_label)
        
        # 时序分析功能
        temporal_section = QWidget()
        temporal_layout = QVBoxLayout(temporal_section)
        temporal_layout.setContentsMargins(0, 0, 0, 0)
        
        temporal_title = QLabel("时序分析增强/Temporal Analysis Enhancement")
        temporal_title.setFont(QFont("Arial", 14, QFont.Bold))
        temporal_title.setStyleSheet("color: #2c3e50; padding: 10px; background-color: rgba(255, 255, 255, 180); border-radius: 4px;")
        temporal_layout.addWidget(temporal_title)
        
        # 时序分析开关
        self.temporal_checkbox = QCheckBox("启用时序分析/Enable Temporal Analysis")
        self.temporal_checkbox.setChecked(self.use_temporal_analysis)
        self.temporal_checkbox.toggled.connect(self.toggle_temporal_analysis)
        self.temporal_checkbox.setStyleSheet("""
            QCheckBox {
                color: #ffffff;
                font-size: 14px;
                padding: 10px;
            }
            QCheckBox::indicator {
                width: 18px;
                height: 18px;
            }
            QCheckBox::indicator:unchecked {
                border: 2px solid #ffffff;
                background-color: transparent;
                border-radius: 3px;
            }
            QCheckBox::indicator:checked {
                border: 2px solid #ffffff;
                background-color: #ffffff;
                border-radius: 3px;
            }
        """)
        temporal_layout.addWidget(self.temporal_checkbox)
        
        # 时序分析参数
        params_widget = QWidget()
        params_layout = QHBoxLayout(params_widget)
        
        window_label = QLabel("时间窗口大小/Window Size:")
        window_label.setStyleSheet("color: #ffffff; font-size: 13px;")
        self.window_size_combo = QComboBox()
        self.window_size_combo.addItems(["15帧(0.5秒)", "30帧(1秒)", "60帧(2秒)", "90帧(3秒)"])
        self.window_size_combo.setCurrentText("30帧(1秒)")
        self.window_size_combo.setStyleSheet("""
            QComboBox {
                background-color: rgba(255, 255, 255, 200);
                color: #2c3e50;
                border: 1px solid #2c3e50;
                padding: 5px 10px;
                border-radius: 4px;
                font-size: 13px;
            }
        """)
        
        params_layout.addWidget(window_label)
        params_layout.addWidget(self.window_size_combo)
        params_layout.addStretch()
        
        temporal_layout.addWidget(params_widget)
        advanced_layout.addWidget(temporal_section)
        
        # 动作阈值调整
        threshold_section = QWidget()
        threshold_layout = QVBoxLayout(threshold_section)
        
        threshold_title = QLabel("动作识别阈值/Action Recognition Thresholds")
        threshold_title.setFont(QFont("Arial", 14, QFont.Bold))
        threshold_title.setStyleSheet("color: #2c3e50; padding: 10px; background-color: rgba(255, 255, 255, 180); border-radius: 4px;")
        threshold_layout.addWidget(threshold_title)
        
        threshold_info = QLabel("此功能正在开发中，将支持动态调整各种动作的识别阈值。\nThis feature is under development and will support dynamic adjustment of recognition thresholds for various actions.")
        threshold_info.setStyleSheet("color: #ffffff; padding: 15px; font-style: italic;")
        threshold_info.setWordWrap(True)
        threshold_layout.addWidget(threshold_info)
        
        advanced_layout.addWidget(threshold_section)
        advanced_layout.addStretch()
        
        # 将高级功能标签页添加到标签页控件
        self.tab_widget.addTab(advanced_widget, "高级功能/Advanced")
        
    def setup_extended_tab(self):
        """设置扩展功能标签页"""
        extended_widget = QWidget()
        extended_layout = QVBoxLayout(extended_widget)
        extended_layout.setContentsMargins(20, 20, 20, 20)
        
        # 扩展功能说明
        info_label = QLabel("扩展功能/Extended Functions")
        info_label.setFont(QFont("Arial", 16, QFont.Bold))
        info_label.setAlignment(Qt.AlignCenter)
        info_label.setStyleSheet("color: #ffffff; padding: 20px; background-color: rgba(255, 255, 255, 120); border-radius: 4px; margin-bottom: 20px;")
        extended_layout.addWidget(info_label)
        
        # 数据分析功能
        analysis_section = QWidget()
        analysis_layout = QVBoxLayout(analysis_section)
        
        analysis_title = QLabel("数据分析与可视化/Data Analysis & Visualization")
        analysis_title.setFont(QFont("Arial", 14, QFont.Bold))
        analysis_title.setStyleSheet("color: #2c3e50; padding: 10px; background-color: rgba(255, 255, 255, 180); border-radius: 4px;")
        analysis_layout.addWidget(analysis_title)
        
        features_list = [
            "• 球员动作热力图分析/Player Action Heatmap Analysis",
            "• 团队协作模式识别/Team Coordination Pattern Recognition", 
            "• 比赛节奏分析/Game Rhythm Analysis",
            "• 动作效率统计/Action Efficiency Statistics",
            "• 3D可视化展示/3D Visualization Display"
        ]
        
        for feature in features_list:
            feature_label = QLabel(feature)
            feature_label.setStyleSheet("color: #ffffff; padding: 5px 15px; font-size: 13px;")
            analysis_layout.addWidget(feature_label)
        
        extended_layout.addWidget(analysis_section)
        
        # AI训练功能
        ai_section = QWidget()
        ai_layout = QVBoxLayout(ai_section)
        
        ai_title = QLabel("AI模型训练/AI Model Training")
        ai_title.setFont(QFont("Arial", 14, QFont.Bold))
        ai_title.setStyleSheet("color: #2c3e50; padding: 10px; background-color: rgba(255, 255, 255, 180); border-radius: 4px;")
        ai_layout.addWidget(ai_title)
        
        ai_features = [
            "• 自定义动作标注/Custom Action Annotation",
            "• 模型微调训练/Model Fine-tuning",
            "• 数据集管理/Dataset Management",
            "• 模型性能评估/Model Performance Evaluation"
        ]
        
        for feature in ai_features:
            feature_label = QLabel(feature)
            feature_label.setStyleSheet("color: #ffffff; padding: 5px 15px; font-size: 13px;")
            ai_layout.addWidget(feature_label)
        
        extended_layout.addWidget(ai_section)
        
        # 开发状态提示
        status_label = QLabel("以上功能正在积极开发中，敬请期待！\nThe above features are under active development, stay tuned!")
        status_label.setAlignment(Qt.AlignCenter)
        status_label.setStyleSheet("color: #ff8f00; padding: 20px; background-color: rgba(255, 255, 255, 180); border-radius: 4px; font-weight: bold; border: 2px dashed #ff8f00;")
        extended_layout.addWidget(status_label)
        
        extended_layout.addStretch()
        
        # 将扩展功能标签页添加到标签页控件
        self.tab_widget.addTab(extended_widget, "扩展功能/Extended")
    
    def toggle_temporal_analysis(self, checked):
        """切换时序分析功能"""
        self.use_temporal_analysis = checked
        if checked:
            print("时序分析已启用/Temporal Analysis Enabled")
        else:
            print("时序分析已禁用/Temporal Analysis Disabled")
            # 重置时序分析器
            reset_temporal_analyzers()
        
    def load_video(self):
        if self.video_source == "file":
            file_name, _ = QFileDialog.getOpenFileName(self, "选择视频文件/Select Video File", "", "Video Files (*.mp4 *.avi)")
            if file_name:
                self.cap = cv2.VideoCapture(file_name)
                self.current_video_path = file_name  # 保存当前视频路径
                
                # 获取视频信息并初始化时间轴
                self.video_fps = self.cap.get(cv2.CAP_PROP_FPS) or 30.0
                total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
                self.video_duration = total_frames / self.video_fps
                self.timeline_widget.set_video_info(self.video_duration, self.video_fps)
                self.timeline_widget.clear_events()  # 清空之前的事件标记
                
                self.init_log_file(os.path.basename(file_name))
                
                # 尝试加载该视频的历史线段
                video_basename = os.path.basename(file_name)
                self.court_lines = self.load_lines_for_video(video_basename)
                
                # 如果没有历史线段，询问是否要标注
                if not self.court_lines:
                    if QMessageBox.question(self, "线段标注/Line Annotation", 
                                            "未找到该视频的线段数据，是否要标注场地线段？\nNo line data found for this video, do you want to annotate court lines?", 
                                            QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                        self.hide()
                        self.court_lines = annotate_court_lines(file_name)
                        if self.court_lines and len(self.court_lines) > 0:
                            num_lines = len(self.court_lines)
                            QMessageBox.information(self, "标注完成/Annotation Complete", 
                                                    f"成功标注了{num_lines}条线段！\nSuccessfully annotated {num_lines} lines!")
                            self.save_lines_for_video(video_basename, self.court_lines)
                            if self.log_file is not None and not self.log_file.closed:
                                self.log_file.write(f"标注的线段信息:\n")
                                for i, line in enumerate(self.court_lines):
                                    start_point, end_point = line
                                    self.log_file.write(f"线段 {i+1}: 从 {start_point} 到 {end_point}\n")
                        self.show()
                else:
                    QMessageBox.information(self, "线段数据/Line Data", 
                                        f"已加载{len(self.court_lines)}条历史线段数据。\nLoaded {len(self.court_lines)} historical line segments.")
        else:
            # 连接相机
            try:
                self.cap = cv2.VideoCapture(0)
                self.current_video_path = "camera_feed"  # 为相机设置标识符
                if not self.cap.isOpened():
                    print("无法连接到相机/Cannot connect to camera")
                    QMessageBox.critical(self, "错误/Error", "无法连接到相机/Cannot connect to camera")
                    return
                self.init_log_file("camera_feed")
                # 加载或标注线段
                self.court_lines = self.load_lines_for_video("camera_feed")
                if not self.court_lines:
                    if QMessageBox.question(self, "线段标注/Line Annotation", 
                                            "未找到相机的线段数据，是否要标注场地线段？\nNo line data found for camera, do you want to annotate court lines?", 
                                            QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                        self.hide()
                        # 采集一帧用于标注
                        ret, frame = self.cap.read()
                        if not ret:
                            QMessageBox.critical(self, "错误/Error", "无法从相机读取画面，无法标注线段！")
                            self.show()
                            return
                        # 临时保存一帧图片用于标注
                        temp_img_path = os.path.join(get_app_root(), "camera_temp_frame.jpg")
                        cv2.imwrite(temp_img_path, frame)
                        self.court_lines = annotate_court_lines(temp_img_path)
                        os.remove(temp_img_path)
                        if self.court_lines and len(self.court_lines) > 0:
                            QMessageBox.information(self, "标注完成/Annotation Complete", 
                                                    f"成功标注了{len(self.court_lines)}条线段！\nSuccessfully annotated {len(self.court_lines)} lines!")
                            self.save_lines_for_video("camera_feed", self.court_lines)
                            if self.log_file is not None and not self.log_file.closed:
                                self.log_file.write(f"标注的线段信息:\n")
                                for i, line in enumerate(self.court_lines):
                                    start_point, end_point = line
                                    self.log_file.write(f"线段 {i+1}: 从 {start_point} 到 {end_point}\n")
                        self.show()
                else:
                    QMessageBox.information(self, "线段数据/Line Data", 
                                        f"已加载{len(self.court_lines)}条相机的历史线段数据。\nLoaded {len(self.court_lines)} historical line segments for camera.")
            except Exception as e:
                print(f"相机连接错误/Camera connection error: {e}")
                QMessageBox.critical(self, "错误/Error", f"相机连接错误/Camera connection error: {str(e)}")
    
    def init_log_file(self, source_name):
        if not os.path.exists("logs"):
            os.makedirs("logs")
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_filename = f"logs/{source_name}_{timestamp}.txt"
        self.log_file = open(log_filename, "w", encoding="utf-8")
        self.log_file.write(f"分析源/Source: {source_name}\n")
        self.log_file.write(f"开始时间/Start Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        self.log_file.write("="*50 + "\n\n")
        self.frame_count = 0
            
    def start_analysis(self):
        if self.cap is not None and self.cap.isOpened():
            if not self.timer.isActive():
                # 重置时序分析器
                reset_temporal_analyzers()
                self.init_video_writer()  # 新增：初始化视频写入器
                self.timer.start(33) # ~30fps
                
                # 更新按钮状态
                self.start_button.setEnabled(False)
                self.pause_button.setEnabled(True)
                self.pause_button.setText("暂停/Pause")
        else:
            print("请先加载视频或连接相机/Please load video or connect camera first.")
            # 可以添加 QMessageBox 提示用户
    
    def pause_analysis(self):
        """暂停/恢复分析"""
        if self.timer.isActive():
            # 当前正在播放，暂停
            self.timer.stop()
            self.pause_button.setText("继续/Resume")
        else:
            # 当前已暂停，恢复播放
            if self.cap is not None and self.cap.isOpened():
                self.timer.start(33)
                self.pause_button.setText("暂停/Pause")
            
    def stop_analysis(self):
        self.timer.stop()
        
        # 重置按钮状态
        self.start_button.setEnabled(True)
        self.pause_button.setEnabled(False)
        self.pause_button.setText("暂停/Pause")
        
        # 新增：释放视频写入器
        if self.video_writer is not None:
            self.video_writer.release()
            self.video_writer = None
            print(f"处理后视频已保存到: {self.output_video_path}")
        if self.log_file is not None and not self.log_file.closed:
            self.log_file.write(f"\n\n结束时间/End Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            self.log_file.write(f"总计分析帧数/Total Frames Analyzed: {self.frame_count}\n")
            self.log_file.close()
            print(f"分析日志已保存/Analysis log saved.")
        
    def update_frame(self):
        if self.cap is None or not self.cap.isOpened():
            self.stop_analysis()
            return
            
        ret, frame = self.cap.read()
        if not ret:
            self.stop_analysis() 
            print("视频结束或读取错误/Video ended or read error.")
            return
        
        self.frame_count += 1
        
        # 检查是否需要跳转到指定帧
        if self.seek_frame >= 0:
            if self.cap.set(cv2.CAP_PROP_POS_FRAMES, self.seek_frame):
                self.frame_count = self.seek_frame
                ret, frame = self.cap.read()
                if not ret:
                    self.stop_analysis()
                    return
            self.seek_frame = -1  # 重置跳转标志
        
        # 更新时间轴显示当前时间
        self.timeline_widget.update_current_time(self.frame_count)
        
        original_frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        log_content = f"帧 {self.frame_count}:\n"
        
        # --- Processing --- 
        detect_results = self.detect_model(frame)
        
        # 使用检测配置模块过滤检测结果
        detect_boxes, detect_classes, detect_confidences = filter_detections(detect_results)
        
        # --- Map Detected Class Indices to Player IDs and Update Headers --- 
        mapped_player_ids = []
        for cls_idx in detect_classes:
            # Use imported mapping
            player_id = player_id_mapping.get(int(cls_idx)) 
            if player_id is not None: 
                mapped_player_ids.append(player_id)

        unique_mapped_ids = sorted(list(set(mapped_player_ids)))
        self.update_table_headers(unique_mapped_ids) 
        
        # --- Pose Estimation and Filtering --- 
        pose_results = self.pose_model(frame)
        filtered_pose_results, keep_indices, matched_classes = self.filter_pose_results(pose_results, detect_boxes)
        
        # --- Process Each Detected Person for Pose and Action --- 
        current_frame_actions = {} 
        pose_frame_for_combined = frame.copy()
        
        if pose_results[0].boxes is not None and len(keep_indices) > 0:
            all_keypoints_data = None
            if hasattr(pose_results[0], 'keypoints') and pose_results[0].keypoints is not None:
                 all_keypoints_data = pose_results[0].keypoints.xy.cpu().numpy() 
            
            for i, original_pose_index in enumerate(keep_indices):
                # Get the bounding box for drawing
                box = pose_results[0].boxes.xyxy.cpu().numpy()[original_pose_index]
                x1, y1, x2, y2 = box.astype(int)
                cv2.rectangle(pose_frame_for_combined, (x1, y1), (x2, y2), (0, 255, 0), 2)

                person_keypoints = None
                if all_keypoints_data is not None and original_pose_index < len(all_keypoints_data):
                    person_keypoints = all_keypoints_data[original_pose_index]
                    # Draw keypoints and connections on the combined frame
                    try:
                        # 定义人体骨骼连接关系
                        # 关键点索引对应：0-nose, 1-left_eye, 2-right_eye, 3-left_ear, 4-right_ear, 
                        # 5-left_shoulder, 6-right_shoulder, 7-left_elbow, 8-right_elbow, 9-left_wrist, 10-right_wrist,
                        # 11-left_hip, 12-right_hip, 13-left_knee, 14-right_knee, 15-left_ankle, 16-right_ankle
                        
                        # 定义骨架连接，分为面部和身体部分
                        face_skeleton = [
                            [0, 1], [0, 2], [1, 3], [2, 4]
                        ]
                        
                        body_skeleton = [
                            # 躯干
                            [5, 6], [5, 11], [6, 12], [11, 12],
                            # 左臂
                            [5, 7], [7, 9],
                            # 右臂
                            [6, 8], [8, 10],
                            # 左腿
                            [11, 13], [13, 15],
                            # 右腿
                            [12, 14], [14, 16]
                        ]
                        
                        # 检查面部关键点是否有效
                        face_valid = True
                        for kp_idx in [0, 1, 2, 3, 4]:  # 面部关键点
                            if (np.isnan(person_keypoints[kp_idx, 0]) or 
                                np.isnan(person_keypoints[kp_idx, 1]) or
                                person_keypoints[kp_idx, 0] <= 0 or 
                                person_keypoints[kp_idx, 1] <= 0):
                                face_valid = False
                        
                        # 画身体骨骼连接
                        for pair in body_skeleton:
                            if (not np.isnan(person_keypoints[pair[0], 0]) and 
                                not np.isnan(person_keypoints[pair[0], 1]) and
                                not np.isnan(person_keypoints[pair[1], 0]) and
                                not np.isnan(person_keypoints[pair[1], 1]) and
                                person_keypoints[pair[0], 0] > 0 and
                                person_keypoints[pair[0], 1] > 0 and
                                person_keypoints[pair[1], 0] > 0 and
                                person_keypoints[pair[1], 1] > 0):
                                
                                pt1 = (int(person_keypoints[pair[0], 0]), int(person_keypoints[pair[0], 1]))
                                pt2 = (int(person_keypoints[pair[1], 0]), int(person_keypoints[pair[1], 1]))
                                cv2.line(pose_frame_for_combined, pt1, pt2, (0, 255, 255), 2)
                        
                        # 只有当面部有效时才画面部连接
                        if face_valid:
                            for pair in face_skeleton:
                                if (not np.isnan(person_keypoints[pair[0], 0]) and 
                                    not np.isnan(person_keypoints[pair[0], 1]) and
                                    not np.isnan(person_keypoints[pair[1], 0]) and
                                    not np.isnan(person_keypoints[pair[1], 1]) and
                                    person_keypoints[pair[0], 0] > 0 and
                                    person_keypoints[pair[0], 1] > 0 and
                                    person_keypoints[pair[1], 0] > 0 and
                                    person_keypoints[pair[1], 1] > 0):
                                    
                                    pt1 = (int(person_keypoints[pair[0], 0]), int(person_keypoints[pair[0], 1]))
                                    pt2 = (int(person_keypoints[pair[1], 0]), int(person_keypoints[pair[1], 1]))
                                    
                                    # 额外检查点是否在合理范围内（防止连到左上角）
                                    if (pt1[0] > 10 and pt1[1] > 10 and pt2[0] > 10 and pt2[1] > 10 and
                                        abs(pt1[0] - pt2[0]) < 100 and abs(pt1[1] - pt2[1]) < 100):
                                        cv2.line(pose_frame_for_combined, pt1, pt2, (0, 255, 255), 2)
                        
                        # 画关键点
                        for kp_idx, kp in enumerate(person_keypoints):
                            if (not np.isnan(kp[0]) and not np.isnan(kp[1]) and 
                                kp[0] > 0 and kp[1] > 0 and
                                # 对于面部点额外检查
                                (kp_idx > 4 or (kp_idx <= 4 and kp[0] > 10 and kp[1] > 10))):
                                cv2.circle(pose_frame_for_combined, (int(kp[0]), int(kp[1])), 4, (0, 0, 255), -1)
                    except Exception as e:
                        print(f"Error drawing pose skeleton: {e}")

                # --- Associate Action with Player ID and Recognize Action --- 
                if i in matched_classes: 
                    detect_idx = matched_classes[i]
                    if detect_idx < len(detect_classes):
                        raw_cls_idx = detect_classes[detect_idx]
                        # Use imported mapping
                        player_id = player_id_mapping.get(int(raw_cls_idx))
                        if player_id is not None:
                            # 根据设置选择识别方法
                            if self.use_temporal_analysis:
                                # 使用时序分析增强的动作识别
                                action_result = recognize_action_with_temporal(
                                    person_keypoints, player_id, (x1, y1, x2, y2)
                                )
                                action_label = action_result[0]
                                confidence = action_result[1]
                                temporal_features = action_result[2]
                                
                                # 在可视化中显示置信度
                                display_text = f"ID:{player_id} {action_label} ({confidence:.2f})"
                            else:
                                # 使用传统几何规则识别
                                action_label = recognize_action(person_keypoints)
                                display_text = f"ID:{player_id} {action_label}"
                            
                            # 记录当前帧的动作信息
                            current_frame_actions[player_id] = action_label
                            
                            # 新增方法：判断点在哪个区域
                            area_num = self.get_player_area(person_keypoints)
                            
                            # 添加区域信息和在图像上显示
                            if area_num > 0:
                                # 存储球员所在区域和动作
                                if self.frame_count not in self.player_area_actions:
                                    self.player_area_actions[self.frame_count] = {}
                                self.player_area_actions[self.frame_count][player_id] = (area_num, action_label)
                                
                                # 在图像上显示球员ID、区域和动作
                                info_text = f"ID:{player_id} A:{area_num} {action_label}"
                                cv2.putText(pose_frame_for_combined, info_text, 
                                           (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 0, 255), 2)
                                
                                # 更新区域动作统计（仅用于时间轴事件标记）
                                if action_label != "Unknown":
                                    current_time = self.frame_count / self.cap.get(cv2.CAP_PROP_FPS) if self.cap is not None else 0
                                    if not self.action_cooldown_manager.is_action_in_cooldown(player_id, action_label, current_time):
                                        if action_label not in self.area_action_stats[area_num]:
                                            self.area_action_stats[area_num][action_label] = 0
                                        self.area_action_stats[area_num][action_label] += 1
                                        
                                        # 添加事件到时间轴
                                        self.timeline_widget.add_event(self.frame_count, player_id, action_label)
                else:
                    # 如果没有关联到球员ID，使用传统方法识别动作
                    action_label = recognize_action(person_keypoints)

        self.frame_actions[self.frame_count] = current_frame_actions
        
        # --- Update the action statistics in the table ---
        self.update_action_statistics()
        
        print(f"Frame {self.frame_count} Actions: {current_frame_actions}")

        # --- 在视频帧上绘制已保存的线段 ---
        if self.court_lines and len(self.court_lines) > 0:
            # 存储所有三等分点的字典，格式为 {(线段编号.点编号): (x, y)}
            third_points = {}
            
            # 第一步：计算并绘制所有线段及其三等分点
            for i, line in enumerate(self.court_lines):
                start_point, end_point = line
                # 使用加粗的红色线条
                cv2.line(pose_frame_for_combined, start_point, end_point, (0, 0, 255), 3)  # 红色加粗线段
                # 在线段两端绘制点
                cv2.circle(pose_frame_for_combined, start_point, 5, (255, 0, 0), -1)  # 蓝色起点
                cv2.circle(pose_frame_for_combined, end_point, 5, (255, 255, 0), -1)  # 黄色终点
                
                # 添加线段编号，方便用户识别
                mid_point = ((start_point[0] + end_point[0]) // 2, (start_point[1] + end_point[1]) // 2)
                cv2.putText(pose_frame_for_combined, f"{i+1}", mid_point, cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
                
                # 计算线段的三等分点
                first_third = (
                    int(start_point[0] + (end_point[0] - start_point[0]) / 3),
                    int(start_point[1] + (end_point[1] - start_point[1]) / 3)
                )
                second_third = (
                    int(start_point[0] + 2 * (end_point[0] - start_point[0]) / 3),
                    int(start_point[1] + 2 * (end_point[1] - start_point[1]) / 3)
                )
                
                # 将三等分点保存到字典中
                third_points[f"{i+1}.1"] = first_third
                third_points[f"{i+1}.2"] = second_third
                
                # 绘制三等分点，使用绿色圆点
                cv2.circle(pose_frame_for_combined, first_third, 4, (0, 255, 0), -1)  # 第一个三等分点
                cv2.circle(pose_frame_for_combined, second_third, 4, (0, 255, 0), -1)  # 第二个三等分点
                
                # 添加三等分点编号，字体小一点，颜色为白色
                cv2.putText(pose_frame_for_combined, f"{i+1}.1", 
                            (first_third[0] + 5, first_third[1] + 5), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
                cv2.putText(pose_frame_for_combined, f"{i+1}.2", 
                            (second_third[0] + 5, second_third[1] + 5), 
                            cv2.FONT_HERSHEY_SIMPLEX, 0.6, (255, 255, 255), 1)
            
            # 第二步：根据要求连接特定的三等分点对
            # 要连接的三等分点对：2.1与4.2, 2.2与4.1, 1.2与3.1, 1.1与3.2
            for idx, (start_key, end_key) in enumerate(self.connection_pairs):
                # 检查这两个点是否存在
                if start_key in third_points and end_key in third_points:
                    start_point = third_points[start_key]
                    end_point = third_points[end_key]
                    
                    # 使用黑色虚线连接这两个点
                    # 首先绘制较粗的白色虚线作为背景/描边
                    for i in range(0, 100, 10):
                        pt1_x = int(start_point[0] + (end_point[0] - start_point[0]) * i / 100)
                        pt1_y = int(start_point[1] + (end_point[1] - start_point[1]) * i / 100)
                        pt2_x = int(start_point[0] + (end_point[0] - start_point[0]) * (i + 5) / 100)
                        pt2_y = int(start_point[1] + (end_point[1] - start_point[1]) * (i + 5) / 100)
                        # 先绘制白色背景
                        cv2.line(pose_frame_for_combined, (pt1_x, pt1_y), (pt2_x, pt2_y), (255, 255, 255), 4)
                        # 再绘制黑色线条
                        cv2.line(pose_frame_for_combined, (pt1_x, pt1_y), (pt2_x, pt2_y), (0, 0, 0), 2)
                    
                    # 在连接线上添加编号
                    conn_mid_point = (
                        (start_point[0] + end_point[0]) // 2,
                        (start_point[1] + end_point[1]) // 2
                    )
                    # 添加文字背景
                    cv2.putText(pose_frame_for_combined, f"C{idx+1}", 
                                (conn_mid_point[0]+2, conn_mid_point[1]+2), 
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)  # 白色描边
                    # 添加黑色文字
                    cv2.putText(pose_frame_for_combined, f"C{idx+1}", 
                                conn_mid_point, 
                                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (0, 0, 0), 2)  # 黑色文字

        pose_frame_for_combined_rgb = cv2.cvtColor(pose_frame_for_combined, cv2.COLOR_BGR2RGB)

        # --- Logging --- 
        log_content += "姿态估计包围框信息:\n"
        if pose_results[0].boxes is not None and len(keep_indices) > 0:
            pose_boxes = pose_results[0].boxes.xyxy.cpu().numpy()[keep_indices]
            for i, box in enumerate(pose_boxes):
                x1, y1, x2, y2 = box
                box_info = f"人物 {i}: 坐标=[{x1:.1f}, {y1:.1f}, {x2:.1f}, {y2:.1f}], 宽度={x2-x1:.1f}, 高度={y2-y1:.1f}"
                log_content += box_info + "\n"
        
        log_content += "\n目标检测包围框信息:\n"
        if detect_results[0].boxes is not None:
            # 使用过滤后的检测结果记录日志
            for i, (box, cls_id_raw, conf) in enumerate(zip(detect_boxes, detect_classes, detect_confidences)):
                x1, y1, x2, y2 = box
                cls_name_raw = f"类别索引{int(cls_id_raw)}"
                # Use imported mapping
                mapped_id = player_id_mapping.get(int(cls_id_raw), "N/A") 
                box_info = f"{cls_name_raw} (球员ID: {mapped_id}) {i}: 坐标=[{x1:.1f}, {y1:.1f}, {x2:.1f}, {y2:.1f}], 宽度={x2-x1:.1f}, 高度={y2-y1:.1f}, 置信度={conf:.3f}"
                log_content += box_info + "\n"

        log_content += "\n位姿信息和目标检测类别关联信息:\n"
        if len(keep_indices) > 0 and len(matched_classes) > 0:
            for pose_idx, detect_idx in matched_classes.items():
                actual_pose_index_in_original = keep_indices[pose_idx]
                if actual_pose_index_in_original < len(pose_results[0].boxes) and detect_idx < len(detect_classes):
                    pose_box = pose_results[0].boxes.xyxy.cpu().numpy()[actual_pose_index_in_original]
                    detect_cls_raw = detect_classes[detect_idx] 
                    # Use imported mapping
                    mapped_id = player_id_mapping.get(int(detect_cls_raw), "N/A") 
                    cls_name = f"球员ID {mapped_id} (类别索引 {int(detect_cls_raw)})"
                    keypoints_info = ""
                    if hasattr(pose_results[0], 'keypoints') and pose_results[0].keypoints is not None:
                        try:
                            keypoints = pose_results[0].keypoints.xy.cpu().numpy()[actual_pose_index_in_original]
                            keypoints_info = f", 关键点数量: {len(keypoints)}"
                        except (IndexError, AttributeError):
                             pass
                    x1, y1, x2, y2 = pose_box
                    association_info = f"姿态ID {pose_idx} (原{actual_pose_index_in_original}) 关联到 {cls_name}: 坐标=[{x1:.1f}, {y1:.1f}, {x2:.1f}, {y2:.1f}]{keypoints_info}"
                    log_content += association_info + "\n"
        
        if self.log_file is not None and not self.log_file.closed:
            self.log_file.write(log_content)
            self.log_file.write("\n" + "-"*30 + "\n\n")
        
        # --- Create Combined Frame --- 
        combined_frame_rgb = self.create_combined_frame(pose_frame_for_combined_rgb, detect_results[0]) 
    
        # --- Display Frames --- 
        self.display_frame(original_frame_rgb, self.original_label)
        self.display_frame(combined_frame_rgb, self.combined_label)
        # 新增：自动保存处理后视频帧
        if self.video_writer is not None:
            out_frame = cv2.cvtColor(combined_frame_rgb, cv2.COLOR_RGB2BGR)
            self.video_writer.write(out_frame)

        # 计算区域中心点并添加区域编号
        if len(self.court_lines) == 4:  # 确保是完整的场地
            # 计算场地四个角点坐标
            top_left = self.court_lines[0][0]  # 线段1起点
            top_right = self.court_lines[2][0]  # 线段3起点
            bottom_right = self.court_lines[2][1]  # 线段3终点
            bottom_left = self.court_lines[0][1]  # 线段1终点
            
            # 计算每个区域的中心点坐标
            width = (top_right[0] - top_left[0]) / 3
            height = (bottom_left[1] - top_left[1]) / 3
            
            for row in range(3):
                for col in range(3):
                    area_num = row * 3 + col + 1  # 区域编号1-9
                    center_x = int(top_left[0] + (col + 0.5) * width)
                    center_y = int(top_left[1] + (row + 0.5) * height)
                    
                    # 添加半透明背景矩形
                    text = f"{area_num}"
                    text_size = cv2.getTextSize(text, cv2.FONT_HERSHEY_SIMPLEX, 2.0, 5)[0]
                    bg_rect_x1 = center_x - text_size[0]//2 - 15
                    bg_rect_y1 = center_y - text_size[1]//2 - 15
                    bg_rect_x2 = center_x + text_size[0]//2 + 15
                    bg_rect_y2 = center_y + text_size[1]//2 + 15
                    
                    # 绘制半透明背景
                    overlay = pose_frame_for_combined.copy()
                    cv2.rectangle(overlay, (bg_rect_x1, bg_rect_y1), (bg_rect_x2, bg_rect_y2), 
                                 (0, 0, 0), -1)
                    # 增加不透明度
                    cv2.addWeighted(overlay, 0.7, pose_frame_for_combined, 0.3, 0, pose_frame_for_combined)
                    
                    # 先绘制白色描边
                    cv2.putText(pose_frame_for_combined, text, 
                               (center_x - text_size[0]//2, center_y + text_size[1]//2), 
                               cv2.FONT_HERSHEY_SIMPLEX, 2.0, (255, 255, 255), 6)
                    # 再绘制亮黄色文字
                    cv2.putText(pose_frame_for_combined, text, 
                               (center_x - text_size[0]//2, center_y + text_size[1]//2), 
                               cv2.FONT_HERSHEY_SIMPLEX, 2.0, (0, 255, 255), 5)
                               
            # 更新显示帧
            pose_frame_for_combined_rgb = cv2.cvtColor(pose_frame_for_combined, cv2.COLOR_BGR2RGB)
            combined_frame_rgb = self.create_combined_frame(pose_frame_for_combined_rgb, detect_results[0])
            self.display_frame(combined_frame_rgb, self.combined_label)

    def process_detection(self, frame, results):
        annotated_frame = results.plot(
            boxes=True,
            line_width=3, # Increased line width
            font_size=10, # Increased font size
            color_mode='instance',
            labels=True,
            probs=False
        )
        return cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2RGB)
        
    def create_combined_frame(self, pose_frame_rgb, detect_result):
        detect_frame_rgb = self.process_detection(cv2.cvtColor(pose_frame_rgb, cv2.COLOR_RGB2BGR), detect_result)
        alpha = 0.6
        combined_frame = cv2.addWeighted(pose_frame_rgb, alpha, detect_frame_rgb, 1-alpha, 0)
        return combined_frame

    def source_changed(self, index):
        if self.cap is not None:
            self.cap.release()
            self.cap = None
        
        self.video_source = "file" if index == 0 else "camera"
        button_text = "加载视频/Load Video" if self.video_source == "file" else "连接相机/Connect Camera"
        self.load_button.setText(button_text)
        # 自动加载相机
        if self.video_source == "camera":
            self.load_video()
            
    def closeEvent(self, event):
        if self.cap is not None:
            self.cap.release()
        self.stop_analysis() # Ensure log file is closed
        event.accept()

    def display_frame(self, frame, label):
        if frame is None or label is None:
            return
        label_width = label.width()
        label_height = label.height()
        if label_width <= 0 or label_height <= 0:
            return

        frame_height, frame_width = frame.shape[:2]
        if frame_height == 0 or frame_width == 0:
             return

        frame_ratio = frame_width / frame_height
        label_ratio = label_width / label_height
        
        if frame_ratio > label_ratio:
            new_width = label_width
            new_height = int(label_width / frame_ratio)
        else:
            new_height = label_height
            new_width = int(label_height * frame_ratio)
        
        if new_width <= 0 or new_height <= 0:
             return # Avoid invalid resize dimensions
             
        resized_frame = cv2.resize(frame, (new_width, new_height), interpolation=cv2.INTER_AREA)
        
        background = np.zeros((label_height, label_width, 3), dtype=np.uint8)
        background[:, :] = (190, 155, 108) # Match blue-gray background RGB(108, 155, 190) in BGR format
        
        y_offset = (label_height - new_height) // 2
        x_offset = (label_width - new_width) // 2
        
        background[y_offset:y_offset+new_height, x_offset:x_offset+new_width] = resized_frame
        
        bytes_per_line = 3 * label_width
        qt_image = QImage(background.data, label_width, label_height, bytes_per_line, QImage.Format_RGB888)
        pixmap = QPixmap.fromImage(qt_image)
        label.setPixmap(pixmap)

    def filter_pose_results(self, pose_results, detect_boxes):
        if not pose_results or not hasattr(pose_results[0], 'boxes') or pose_results[0].boxes is None or len(detect_boxes) == 0:
            # Return empty lists/dicts matching the expected return type
            return pose_results, [], {} 
            
        pose_boxes = pose_results[0].boxes.xyxy.cpu().numpy()
        iou_threshold = 0.3
        keep_mask = np.zeros(len(pose_boxes), dtype=bool)
        matched_classes = {}
        
        for i, pose_box in enumerate(pose_boxes):
            best_iou = 0
            best_match_idx = -1
            for j, detect_box in enumerate(detect_boxes):
                iou = self.calculate_iou(pose_box, detect_box)
                if iou > iou_threshold and iou > best_iou:
                    best_iou = iou
                    best_match_idx = j
                    
            if best_match_idx != -1:
                keep_mask[i] = True
                # Corrected Indentation:
                matched_classes[i] = best_match_idx # Store original index of pose box
        
        keep_indices = np.where(keep_mask)[0]
        
        # Remap matched_classes to use the index within keep_indices
        remapped_matched_classes = {new_idx: matched_classes[orig_idx] 
                                   for new_idx, orig_idx in enumerate(keep_indices) 
                                   if orig_idx in matched_classes} # Ensure orig_idx exists

        if len(keep_indices) == 0:
             return pose_results, [], {}

        return pose_results, keep_indices, remapped_matched_classes
        
    def calculate_iou(self, box1, box2):
        x1_1, y1_1, x2_1, y2_1 = box1
        x1_2, y1_2, x2_2, y2_2 = box2
        x1_i = max(x1_1, x1_2)
        y1_i = max(y1_1, y1_2)
        x2_i = min(x2_1, x2_2)
        y2_i = min(y2_1, y2_2)
        
        if x2_i < x1_i or y2_i < y1_i:
            return 0.0
        
        intersection_area = (x2_i - x1_i) * (y2_i - y1_i)
        box1_area = (x2_1 - x1_1) * (y2_1 - y1_1)
        box2_area = (x2_2 - x1_2) * (y2_2 - y1_2)
        union_area = box1_area + box2_area - intersection_area
        iou = intersection_area / union_area if union_area > 0 else 0.0
        return iou

    # --- 修改表头更新逻辑 ---
    def update_table_headers(self, new_player_ids):
        """更新表格标题，保持已识别球员的固定位置。"""
        # 将新检测到的ID添加到总列表中（如果不存在）
        for player_id in new_player_ids:
            if player_id not in self.all_detected_player_ids:
                self.all_detected_player_ids.append(player_id)
        
        # 如果我们尚未填充表格标题，或检测到了新ID
        if not self.current_displayed_player_ids or set(new_player_ids) - set(self.current_displayed_player_ids):
            # 确保最多显示6个球员ID
            ids_to_display = self.all_detected_player_ids[:self.max_players_to_display]
            
            # 如果显示的ID变化了，更新表格
            if set(ids_to_display) != set(self.current_displayed_player_ids):
                self.current_displayed_player_ids = ids_to_display
                
                # 更新表格标题
                max_player_cols = self.table_model.columnCount() - 1  # 第一列是动作名称
                
                for i in range(max_player_cols):
                    col_index = i + 1  # 从第二列开始（索引1）
                    header_text = str(ids_to_display[i]) if i < len(ids_to_display) else "-"
                    header_item = QStandardItem(header_text)
                    header_item.setTextAlignment(Qt.AlignCenter)
                    self.table_model.setHorizontalHeaderItem(col_index, header_item)

    # --- 新方法：更新动作统计 --- 
    def update_action_statistics(self):
        """根据当前帧的动作信息更新表格的动作统计数据"""
        # 获取当前帧中的动作信息
        if self.frame_count in self.frame_actions:
            # 获取当前时间戳（基于帧率计算）
            current_time = self.frame_count / self.cap.get(cv2.CAP_PROP_FPS) if self.cap is not None else 0
            
            # 遍历每个检测到的球员ID及其对应的动作
            for player_id, action_label in self.frame_actions[self.frame_count].items():
                # 跳过Unknown动作
                if action_label == "Unknown":
                    continue
                
                # 检查是否在冷却时间内，只有不在冷却时间内才增加计数
                if self.action_cooldown_manager.is_action_in_cooldown(player_id, action_label, current_time):
                    continue
                
                # 查找表格中对应的列
                col_index = -1
                for i in range(1, self.table_model.columnCount()):
                    header_item = self.table_model.horizontalHeaderItem(i)
                    if header_item and header_item.text() == str(player_id):
                        col_index = i
                        break
                
                # 如果找到了对应的列
                if col_index != -1:
                    # 根据动作名称找到对应的行
                    row_index = -1
                    for i in range(self.table_model.rowCount()):
                        action_item = self.table_model.item(i, 0)
                        if action_item and action_item.text() == action_label:
                            row_index = i
                            break
                    
                    # 如果找到了对应的行
                    if row_index != -1:
                        # 更新表格中的计数
                        current_item = self.table_model.item(row_index, col_index)
                        current_count = 0
                        if current_item and current_item.text():
                            try:
                                current_count = int(current_item.text())
                            except ValueError:
                                current_count = 0
                        
                        # 创建新的单元格项，计数+1
                        new_item = QStandardItem(str(current_count + 1))
                        new_item.setEditable(False)
                        new_item.setTextAlignment(Qt.AlignCenter)
                        new_item.setForeground(QColor("#ccd6f6"))
                        self.table_model.setItem(row_index, col_index, new_item)
                        
                        # 更新该球员该动作的最后执行时间
                        self.action_cooldown_manager.update_player_last_action(player_id, action_label, current_time)
                        
                        print(f"表格统计更新: 球员{player_id} {action_label} 计数: {current_count + 1}")

    def export_report(self):
        """将当前表格数据导出为Excel报表"""
        if self.table_model.rowCount() <= 1:  # 只有标题行
            QMessageBox.warning(self, "警告", "没有数据可导出！\nNo data to export!")
            return
            
        # 获取保存文件路径
        file_path, _ = QFileDialog.getSaveFileName(
            self, 
            "保存报表/Save Report", 
            os.path.join(os.path.expanduser("~"), "Desktop", f"排球分析报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"),
            "Excel Files (*.xlsx)"
        )
        
        if not file_path:
            return  # 用户取消了保存
            
        try:
            # 创建DataFrame
            data = []
            
            # 获取表头
            headers = []
            for i in range(self.table_model.columnCount()):
                header_item = self.table_model.horizontalHeaderItem(i)
                headers.append(header_item.text() if header_item else f"Column {i}")
            
            # 获取表格数据
            for row in range(self.table_model.rowCount()):
                row_data = []
                for col in range(self.table_model.columnCount()):
                    item = self.table_model.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            
            # 创建DataFrame
            df = pd.DataFrame(data, columns=headers)
            
            # 添加时间戳和视频信息
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            video_info = f"视频源: {os.path.basename(self.video_source) if self.video_source else '未加载视频'}"
            frame_info = f"分析帧数: {self.frame_count}"
            
            # 创建Excel写入器
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # 写入主数据表
                df.to_excel(writer, sheet_name='动作统计', index=False)
                
                # 获取工作簿和工作表
                workbook = writer.book
                worksheet = writer.sheets['动作统计']
                
                # 添加标题和元数据
                worksheet.insert_rows(0, 3)
                worksheet['A1'] = "排球视频分析报表"
                worksheet['A1'].font = openpyxl.styles.Font(size=14, bold=True)
                worksheet['A2'] = timestamp
                worksheet['A3'] = video_info
                worksheet['A4'] = frame_info
                
                # 调整列宽
                for col in range(1, len(headers) + 1):
                    column_letter = openpyxl.utils.get_column_letter(col)
                    worksheet.column_dimensions[column_letter].width = 15
                
                # 创建第二个工作表 - 帧动作详情
                if self.frame_actions:
                    frame_data = []
                    for frame_num, actions in self.frame_actions.items():
                        for player_id, action in actions.items():
                            frame_data.append({
                                '帧号': frame_num,
                                '球员ID': player_id,
                                '动作': action,
                                '时间(秒)': frame_num / self.cap.get(cv2.CAP_PROP_FPS) if self.cap else 0
                            })
                    
                    if frame_data:
                        frame_df = pd.DataFrame(frame_data)
                        frame_df.to_excel(writer, sheet_name='帧动作详情', index=False)
                        
                        # 调整列宽
                        frame_worksheet = writer.sheets['帧动作详情']
                        for col in range(1, len(frame_df.columns) + 1):
                            column_letter = openpyxl.utils.get_column_letter(col)
                            frame_worksheet.column_dimensions[column_letter].width = 15
                
                # --- 新增：如果有线段数据，添加线段信息工作表 ---
                if self.court_lines and len(self.court_lines) > 0:
                    lines_data = []
                    
                    # 创建三等分点字典用于后续添加连接线
                    third_points = {}
                    
                    for i, line in enumerate(self.court_lines):
                        start_point, end_point = line
                        length = np.sqrt((end_point[0]-start_point[0])**2 + (end_point[1]-start_point[1])**2)
                        angle = np.arctan2(end_point[1]-start_point[1], end_point[0]-start_point[0]) * 180 / np.pi
                        
                        # 计算三等分点
                        first_third = (
                            int(start_point[0] + (end_point[0] - start_point[0]) / 3),
                            int(start_point[1] + (end_point[1] - start_point[1]) / 3)
                        )
                        second_third = (
                            int(start_point[0] + 2 * (end_point[0] - start_point[0]) / 3),
                            int(start_point[1] + 2 * (end_point[1] - start_point[1]) / 3)
                        )
                        
                        # 保存到字典
                        third_points[f"{i+1}.1"] = first_third
                        third_points[f"{i+1}.2"] = second_third
                        
                        lines_data.append({
                            '线段编号': i+1,
                            '线段类型': '原始线段',
                            '起点X': start_point[0],
                            '起点Y': start_point[1],
                            '终点X': end_point[0],
                            '终点Y': end_point[1],
                            '三等分点1X': first_third[0],
                            '三等分点1Y': first_third[1],
                            '三等分点2X': second_third[0],
                            '三等分点2Y': second_third[1],
                            '长度(像素)': round(length, 2),
                            '角度(度)': round(angle, 2)
                        })
                    
                    # 添加连接线数据
                    for idx, (start_key, end_key) in enumerate(self.connection_pairs):
                        if start_key in third_points and end_key in third_points:
                            start_point = third_points[start_key]
                            end_point = third_points[end_key]
                            
                            length = np.sqrt((end_point[0]-start_point[0])**2 + (end_point[1]-start_point[1])**2)
                            angle = np.arctan2(end_point[1]-start_point[1], end_point[0]-start_point[0]) * 180 / np.pi
                            
                            lines_data.append({
                                '线段编号': f"连接线{idx+1}",
                                '线段类型': f'三等分点连接 ({start_key}-{end_key})',
                                '起点X': start_point[0],
                                '起点Y': start_point[1],
                                '终点X': end_point[0],
                                '终点Y': end_point[1],
                                '三等分点1X': None,
                                '三等分点1Y': None,
                                '三等分点2X': None,
                                '三等分点2Y': None,
                                '长度(像素)': round(length, 2),
                                '角度(度)': round(angle, 2)
                            })
                
                if lines_data:
                    lines_df = pd.DataFrame(lines_data)
                    lines_df.to_excel(writer, sheet_name='场地线段数据', index=False)
                    
                    # 调整列宽
                    lines_worksheet = writer.sheets['场地线段数据']
                    for col in range(1, len(lines_df.columns) + 1):
                        column_letter = openpyxl.utils.get_column_letter(col)
                        lines_worksheet.column_dimensions[column_letter].width = 15

                # --- 添加区域动作分析工作表 ---
                if self.player_area_actions:
                    # 准备区域动作数据
                    area_player_data = []
                    for frame_num, player_info in self.player_area_actions.items():
                        for player_id, (area_num, action) in player_info.items():
                            area_player_data.append({
                                '帧号': frame_num,
                                '球员ID': player_id,
                                '区域': area_num,
                                '动作': action,
                                '时间(秒)': frame_num / self.cap.get(cv2.CAP_PROP_FPS) if self.cap else 0
                            })
                    
                    if area_player_data:
                        area_df = pd.DataFrame(area_player_data)
                        area_df.to_excel(writer, sheet_name='区域动作分析', index=False)
                        
                        # 调整列宽
                        area_worksheet = writer.sheets['区域动作分析']
                        for col in range(1, len(area_df.columns) + 1):
                            column_letter = openpyxl.utils.get_column_letter(col)
                            area_worksheet.column_dimensions[column_letter].width = 15

                # --- 添加区域统计汇总工作表 ---
                area_stat_data = []
                for area_num, action_counts in self.area_action_stats.items():
                    for action, count in action_counts.items():
                        area_stat_data.append({
                            '区域': area_num,
                            '动作': action,
                            '次数': count
                        })

                if area_stat_data:
                    area_stat_df = pd.DataFrame(area_stat_data)
                    area_stat_df.to_excel(writer, sheet_name='区域动作统计', index=False)
                    
                    # 调整列宽
                    area_stat_worksheet = writer.sheets['区域动作统计']
                    for col in range(1, len(area_stat_df.columns) + 1):
                        column_letter = openpyxl.utils.get_column_letter(col)
                        area_stat_worksheet.column_dimensions[column_letter].width = 15
            
            QMessageBox.information(self, "成功", f"报表已成功导出到:\n{file_path}\n\nReport successfully exported to:\n{file_path}")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出报表时发生错误:\n{str(e)}\n\nError occurred while exporting report:\n{str(e)}")

    # --- 新增：线段历史管理相关方法 ---
    def save_lines_for_video(self, video_name, lines):
        """保存视频的线段数据"""
        if not lines:
            return
            
        # 生成文件名，将视频文件名转换为安全的文件名
        safe_name = ''.join(c if c.isalnum() else '_' for c in video_name)
        file_path = os.path.join(self.lines_history_dir, f"{safe_name}_lines.json")
        
        # 将线段数据转换为可序列化的格式
        serializable_lines = []
        for line in lines:
            start_point, end_point = line
            # 计算三等分点
            first_third = (
                int(start_point[0] + (end_point[0] - start_point[0]) / 3),
                int(start_point[1] + (end_point[1] - start_point[1]) / 3)
            )
            second_third = (
                int(start_point[0] + 2 * (end_point[0] - start_point[0]) / 3),
                int(start_point[1] + 2 * (end_point[1] - start_point[1]) / 3)
            )
            serializable_lines.append({
                "start": [int(start_point[0]), int(start_point[1])],
                "end": [int(end_point[0]), int(end_point[1])],
                "first_third": [first_third[0], first_third[1]],
                "second_third": [second_third[0], second_third[1]]
            })
            
        # 保存为JSON文件
        with open(file_path, 'w') as f:
            json.dump(serializable_lines, f)
        
        print(f"已保存线段数据到 {file_path}")
    
    def load_lines_for_video(self, video_name):
        """加载视频的线段数据"""
        # 生成文件名
        safe_name = ''.join(c if c.isalnum() else '_' for c in video_name)
        file_path = os.path.join(self.lines_history_dir, f"{safe_name}_lines.json")
        
        if not os.path.exists(file_path):
            return []
            
        try:
            with open(file_path, 'r') as f:
                serialized_lines = json.load(f)
                
            # 将序列化数据转换回线段格式
            lines = []
            for line_data in serialized_lines:
                start = tuple(line_data["start"])
                end = tuple(line_data["end"])
                # 注意：我们只需要保存起点和终点，三等分点会在绘制时计算
                lines.append([start, end])
                
            print(f"已加载{len(lines)}条线段数据从 {file_path}")
            return lines
        except Exception as e:
            print(f"加载线段数据失败: {e}")
            return []
    
    def get_all_saved_videos(self):
        """获取所有有保存线段的视频名称"""
        videos = []
        for filename in os.listdir(self.lines_history_dir):
            if filename.endswith("_lines.json"):
                video_name = filename.replace("_lines.json", "")
                # 还原原始文件名（尽量）
                original_name = video_name.replace('_', ' ')
                videos.append((video_name, original_name))
        return videos
        
    def manage_court_lines(self):
        """管理线段历史数据"""
        if not self.current_video_path:
            QMessageBox.warning(self, "警告/Warning", 
                             "请先加载视频文件！\nPlease load a video file first!")
            return
            
        video_basename = os.path.basename(self.current_video_path)
        
        # 创建选项菜单
        options = [
            "重新标注当前视频线段/Re-annotate current video", 
            "从其他视频导入线段/Import lines from other video",
            "删除当前视频线段/Delete current video lines",
            "查看所有已保存线段/View all saved lines",
            "删除指定线段/Delete specific lines"  # 添加新选项
        ]
        
        selected_option, ok = QInputDialog.getItem(
            self, "线段管理/Line Management", 
            "请选择操作/Select operation:", 
            options, 0, False
        )
        
        if not ok:
            return
            
        # 处理选项
        if selected_option == options[0]:  # 重新标注
            if self.cap is None:
                QMessageBox.warning(self, "警告/Warning", 
                                "无法标注：视频未加载或已关闭！\nCannot annotate: Video not loaded or closed!")
                return
                
            # 弹出标注工具
            self.hide()
            new_lines = annotate_court_lines(self.current_video_path)
            if new_lines and len(new_lines) > 0:
                self.court_lines = new_lines
                self.save_lines_for_video(video_basename, self.court_lines)
                QMessageBox.information(self, "标注完成/Annotation Complete", 
                                    f"成功标注了{len(new_lines)}条线段！\nSuccessfully annotated {len(new_lines)} lines!")
            self.show()
            
        elif selected_option == options[1]:  # 从其他视频导入
            saved_videos = self.get_all_saved_videos()
            if not saved_videos:
                QMessageBox.information(self, "信息/Information", 
                                     "没有找到已保存的线段数据！\nNo saved line data found!")
                return
                
            video_names = [f"{original} ({safe})" for safe, original in saved_videos]
            selected_video, ok = QInputDialog.getItem(
                self, "选择视频/Select Video", 
                "选择要导入线段的视频/Select video to import lines from:", 
                video_names, 0, False
            )
            
            if ok and selected_video:
                # 提取安全名称
                import_video_name = selected_video.split("(")[1].split(")")[0]
                imported_lines = self.load_lines_for_video(import_video_name)
                if imported_lines:
                    self.court_lines = imported_lines
                    self.save_lines_for_video(video_basename, self.court_lines)
                    QMessageBox.information(self, "导入成功/Import Successful", 
                                         f"已成功导入{len(imported_lines)}条线段！\nSuccessfully imported {len(imported_lines)} lines!")
        
        elif selected_option == options[2]:  # 删除当前线段
            if QMessageBox.question(self, "确认删除/Confirm Deletion", 
                                "确定要删除当前视频的所有线段数据吗？\nAre you sure you want to delete all line data for the current video?", 
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                safe_name = ''.join(c if c.isalnum() else '_' for c in video_basename)
                file_path = os.path.join(self.lines_history_dir, f"{safe_name}_lines.json")
                
                if os.path.exists(file_path):
                    os.remove(file_path)
                    self.court_lines = []
                    QMessageBox.information(self, "删除成功/Deletion Successful", 
                                         "已成功删除线段数据！\nLine data successfully deleted!")
        
        elif selected_option == options[3]:  # 查看所有已保存
            saved_videos = self.get_all_saved_videos()
            if not saved_videos:
                QMessageBox.information(self, "信息/Information", 
                                     "没有找到已保存的线段数据！\nNo saved line data found!")
                return
                
            info_text = "已保存的线段数据/Saved line data:\n\n"
            for _, original_name in saved_videos:
                info_text += f"- {original_name}\n"
                
            QMessageBox.information(self, "线段数据列表/Line Data List", info_text)
        
        elif selected_option == options[4]:  # 删除指定线段
            if not self.court_lines or len(self.court_lines) == 0:
                QMessageBox.information(self, "信息/Information", 
                                     "当前没有线段数据！\nNo line data available!")
                return
                
            # 打开线段选择对话框
            dialog = LineSelectionDialog(self.court_lines, self)
            if dialog.exec_() == QDialog.Accepted:
                # 获取未被选中的线段（保留的线段）
                kept_lines = []
                for i, line in enumerate(self.court_lines):
                    if not dialog.is_line_selected(i):
                        kept_lines.append(line)
                
                # 更新线段数据
                self.court_lines = kept_lines
                self.save_lines_for_video(video_basename, self.court_lines)
                
                # 显示结果
                QMessageBox.information(self, "删除成功/Deletion Successful", 
                                     f"已删除选中的线段，剩余{len(kept_lines)}条线段。\nSelected lines deleted, {len(kept_lines)} lines remaining.")

    def keyPressEvent(self, event):
        """处理键盘事件"""
        # 添加快捷键Ctrl+D打开删除指定线段对话框
        if event.key() == Qt.Key_D and event.modifiers() == Qt.ControlModifier:
            if self.court_lines and len(self.court_lines) > 0:
                dialog = LineSelectionDialog(self.court_lines, self)
                if dialog.exec_() == QDialog.Accepted:
                    # 获取未被选中的线段（保留的线段）
                    kept_lines = []
                    for i, line in enumerate(self.court_lines):
                        if not dialog.is_line_selected(i):
                            kept_lines.append(line)
                    
                    # 更新线段数据
                    video_basename = os.path.basename(self.current_video_path) if self.current_video_path else "unknown"
                    self.court_lines = kept_lines
                    self.save_lines_for_video(video_basename, self.court_lines)
                    
                    # 显示结果
                    QMessageBox.information(self, "删除成功/Deletion Successful", 
                                        f"已删除选中的线段，剩余{len(kept_lines)}条线段。\nSelected lines deleted, {len(kept_lines)} lines remaining.")
        # 添加数字键选择删除对应编号的线段
        elif event.key() >= Qt.Key_1 and event.key() <= Qt.Key_9 and event.modifiers() == Qt.ControlModifier:
            line_index = event.key() - Qt.Key_1  # 转换为0-8的索引
            if self.court_lines and line_index < len(self.court_lines):
                if QMessageBox.question(self, "确认删除/Confirm Deletion", 
                                    f"确定要删除线段 {line_index+1}？\nAre you sure you want to delete line {line_index+1}?", 
                                    QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
                    # 删除对应线段
                    del self.court_lines[line_index]
                    
                    # 保存更新后的线段数据
                    video_basename = os.path.basename(self.current_video_path) if self.current_video_path else "unknown"
                    self.save_lines_for_video(video_basename, self.court_lines)
                    
                    # 显示结果
                    QMessageBox.information(self, "删除成功/Deletion Successful", 
                                        f"已删除线段 {line_index+1}，剩余{len(self.court_lines)}条线段。\nLine {line_index+1} deleted, {len(self.court_lines)} lines remaining.")

    # 新增方法：判断点在哪个区域
    def get_point_area(self, point):
        """
        判断一个坐标点位于场地的哪个区域(1-9)
        
        参数:
            point: 坐标点 (x, y)
            
        返回:
            area_num: 区域编号(1-9)，如果不在场地内则返回0
        """
        if not self.court_lines or len(self.court_lines) != 4:
            return 0  # 没有完整的场地线段定义
            
        # 获取场地四个角点
        top_left = self.court_lines[0][0]  # 线段1起点
        top_right = self.court_lines[2][0]  # 线段3起点
        bottom_right = self.court_lines[2][1]  # 线段3终点
        bottom_left = self.court_lines[0][1]  # 线段1终点
        
        # 检查点是否在场地范围内
        x, y = point
        if (x < min(top_left[0], bottom_left[0]) or 
            x > max(top_right[0], bottom_right[0]) or
            y < min(top_left[1], top_right[1]) or
            y > max(bottom_left[1], bottom_right[1])):
            return 0  # 点在场地外
        
        # 计算场地的宽度和高度
        width = (top_right[0] - top_left[0]) / 3
        height = (bottom_left[1] - top_left[1]) / 3
        
        # 根据点的相对位置计算区域编号
        col = int((x - top_left[0]) / width)
        row = int((y - top_left[1]) / height)
        
        # 确保行列在合理范围内
        col = max(0, min(col, 2))
        row = max(0, min(row, 2))
        
        # 计算区域编号(1-9)
        area_num = row * 3 + col + 1
        return area_num

    # 新增方法：获取球员所在区域
    def get_player_area(self, person_keypoints):
        """
        根据球员的关键点获取其所在区域
        
        参数:
            person_keypoints: 人体关键点数组 (17, 2)
            
        返回:
            area_num: 区域编号(1-9)，如果不能确定则返回0
        """
        if person_keypoints is None or person_keypoints.shape != (17, 2):
            return 0
        
        # 获取左右脚踝的坐标 (索引15和16)
        left_ankle = person_keypoints[15]
        right_ankle = person_keypoints[16]
        
        left_valid = not np.isnan(left_ankle[0]) and not np.isnan(left_ankle[1]) and left_ankle[0] > 0 and left_ankle[1] > 0
        right_valid = not np.isnan(right_ankle[0]) and not np.isnan(right_ankle[1]) and right_ankle[0] > 0 and right_ankle[1] > 0
        
        # 如果两个脚踝都有效，取中点
        if left_valid and right_valid:
            ankle_x = (left_ankle[0] + right_ankle[0]) / 2
            ankle_y = (left_ankle[1] + right_ankle[1]) / 2
            return self.get_point_area((ankle_x, ankle_y))
        
        # 如果只有一个脚踝有效，使用那个
        elif left_valid:
            return self.get_point_area((left_ankle[0], left_ankle[1]))
        elif right_valid:
            return self.get_point_area((right_ankle[0], right_ankle[1]))
        
        # 如果脚踝都无效，尝试使用髋关节位置
        left_hip = person_keypoints[11]
        right_hip = person_keypoints[12]
        
        left_hip_valid = not np.isnan(left_hip[0]) and not np.isnan(left_hip[1]) and left_hip[0] > 0 and left_hip[1] > 0
        right_hip_valid = not np.isnan(right_hip[0]) and not np.isnan(right_hip[1]) and right_hip[0] > 0 and right_hip[1] > 0
        
        if left_hip_valid and right_hip_valid:
            hip_x = (left_hip[0] + right_hip[0]) / 2
            hip_y = (left_hip[1] + right_hip[1]) / 2
            return self.get_point_area((hip_x, hip_y))
        elif left_hip_valid:
            return self.get_point_area((left_hip[0], left_hip[1]))
        elif right_hip_valid:
            return self.get_point_area((right_hip[0], right_hip[1]))
        
        return 0  # 无法确定位置

    def init_video_writer(self):
        """初始化视频保存器"""
        if self.cap is None:
            return
        width = int(self.cap.get(cv2.CAP_PROP_FRAME_WIDTH))
        height = int(self.cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
        fps = self.cap.get(cv2.CAP_PROP_FPS)
        if fps <= 0:
            fps = 30
        if not os.path.exists("output_videos"):
            os.makedirs("output_videos")
        base_name = os.path.basename(self.current_video_path) if self.current_video_path else "camera_feed"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_name = f"{os.path.splitext(base_name)[0]}_分析_{timestamp}.mp4"
        self.output_video_path = os.path.join("output_videos", out_name)
        fourcc = cv2.VideoWriter_fourcc(*'mp4v')
        self.video_writer = cv2.VideoWriter(self.output_video_path, fourcc, fps, (width, height))
    
    def seek_to_frame(self, target_frame):
        """跳转到指定帧"""
        if self.cap is not None and self.cap.isOpened():
            total_frames = int(self.cap.get(cv2.CAP_PROP_FRAME_COUNT))
            if 0 <= target_frame < total_frames:
                self.seek_frame = target_frame
                print(f"跳转到帧: {target_frame}")

# 添加线段选择对话框类
class LineSelectionDialog(QDialog):
    def __init__(self, lines, parent=None):
        super().__init__(parent)
        self.lines = lines
        self.setWindowTitle("选择要删除的线段/Select Lines to Delete")
        self.resize(600, 400)
        
        # 创建主布局
        layout = QVBoxLayout(self)
        
        # 添加说明标签
        info_label = QLabel("选择要删除的线段（勾选要删除的线段）：\nSelect lines to delete (check the lines you want to delete):")
        layout.addWidget(info_label)
        
        # 创建线段列表
        self.list_widget = QListWidget()
        self.populate_list()
        layout.addWidget(self.list_widget)
        
        # 添加按钮
        button_layout = QHBoxLayout()
        
        select_all_button = QPushButton("全选/Select All")
        select_all_button.clicked.connect(self.select_all)
        button_layout.addWidget(select_all_button)
        
        deselect_all_button = QPushButton("取消全选/Deselect All")
        deselect_all_button.clicked.connect(self.deselect_all)
        button_layout.addWidget(deselect_all_button)
        
        ok_button = QPushButton("确定/OK")
        ok_button.clicked.connect(self.accept)
        button_layout.addWidget(ok_button)
        
        cancel_button = QPushButton("取消/Cancel")
        cancel_button.clicked.connect(self.reject)
        button_layout.addWidget(cancel_button)
        
        layout.addLayout(button_layout)
    
    def populate_list(self):
        """填充线段列表"""
        for i, line in enumerate(self.lines):
            start_point, end_point = line
            line_length = np.sqrt((end_point[0]-start_point[0])**2 + (end_point[1]-start_point[1])**2)
            
            # 创建带复选框的列表项
            item = QListWidgetItem(f"线段 {i+1}：从 ({start_point[0]}, {start_point[1]}) 到 ({end_point[0]}, {end_point[1]})，长度: {line_length:.1f} 像素")
            item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
            item.setCheckState(Qt.Unchecked)
            
            self.list_widget.addItem(item)
    
    def is_line_selected(self, index):
        """检查线段是否被选中删除"""
        if index < self.list_widget.count():
            item = self.list_widget.item(index)
            return item.checkState() == Qt.Checked
        return False
    
    def select_all(self):
        """选择所有线段"""
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(Qt.Checked)
    
    def deselect_all(self):
        """取消选择所有线段"""
        for i in range(self.list_widget.count()):
            self.list_widget.item(i).setCheckState(Qt.Unchecked)

class EventTimelineWidget(QWidget):
    """事件时间轴组件"""
    
    def __init__(self):
        super().__init__()
        self.setFixedHeight(80)
        self.video_duration = 0  # 视频总时长（秒）
        self.current_time = 0    # 当前播放时间（秒）
        self.events = []         # 事件列表: [{'time': float, 'player_id': int, 'action': str, 'frame': int}, ...]
        self.fps = 30           # 帧率
        self.setMouseTracking(True)  # 启用鼠标跟踪
        
        # 样式配置
        self.bg_color = "rgb(108, 155, 190)"
        self.progress_bg_color = "#8a9ba8"
        self.progress_color = "#ffffff"
        self.event_colors = {
            'Serve': '#ff6b6b',      # 红色 - 发球
            'Reception': '#4ecdc4',   # 青色 - 接球  
            'Attack': '#ff8c42',     # 橙色 - 扣球
            'Block': '#6c5ce7',      # 紫色 - 拦网
            'Dig': '#a8e6cf',       # 绿色 - 防守
            'Set': '#ffd93d'        # 黄色 - 传球
        }
        
        self.setStyleSheet(f"background-color: {self.bg_color}; border: 1px solid #8a9ba8; border-radius: 4px;")
    
    def set_video_info(self, duration_seconds, fps):
        """设置视频信息"""
        self.video_duration = duration_seconds
        self.fps = fps
        self.update()
    
    def update_current_time(self, current_frame):
        """更新当前播放时间"""
        if self.fps > 0:
            self.current_time = current_frame / self.fps
            self.update()
    
    def add_event(self, frame_num, player_id, action_label):
        """添加事件标记"""
        if self.fps > 0:
            time_seconds = frame_num / self.fps
            event = {
                'time': time_seconds,
                'player_id': player_id,
                'action': action_label,
                'frame': frame_num
            }
            self.events.append(event)
            self.update()
    
    def clear_events(self):
        """清空所有事件标记"""
        self.events.clear()
        self.update()
    
    def paintEvent(self, event):
        """绘制时间轴"""
        from PyQt5.QtGui import QPainter, QPen, QBrush, QFontMetrics
        
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        
        width = self.width() - 20  # 留出边距
        height = self.height()
        x_offset = 10
        
        # 绘制进度条背景
        progress_y = height // 2 - 3
        progress_height = 6
        painter.fillRect(x_offset, progress_y, width, progress_height, QColor(self.progress_bg_color))
        
        # 绘制当前进度
        if self.video_duration > 0:
            progress_width = int((self.current_time / self.video_duration) * width)
            painter.fillRect(x_offset, progress_y, progress_width, progress_height, QColor(self.progress_color))
        
        # 绘制事件标记
        for event in self.events:
            if self.video_duration > 0:
                x_pos = x_offset + int((event['time'] / self.video_duration) * width)
                
                # 获取动作对应的颜色
                color = self.event_colors.get(event['action'], '#ffffff')
                
                # 绘制事件标记点
                painter.fillRect(x_pos - 3, progress_y - 5, 6, progress_height + 10, QColor(color))
                
                # 绘制小三角形标记
                painter.setBrush(QBrush(QColor(color)))
                painter.setPen(QPen(QColor(color)))
                triangle_points = [
                    (x_pos, progress_y - 8),
                    (x_pos - 4, progress_y - 15),
                    (x_pos + 4, progress_y - 15)
                ]
                from PyQt5.QtGui import QPolygon
                from PyQt5.QtCore import QPoint
                triangle = QPolygon([QPoint(x, y) for x, y in triangle_points])
                painter.drawPolygon(triangle)
        
        # 绘制时间标签
        painter.setPen(QPen(QColor("#ccd6f6")))
        painter.drawText(x_offset, height - 5, f"{int(self.current_time // 60):02d}:{int(self.current_time % 60):02d}")
        
        if self.video_duration > 0:
            total_minutes = int(self.video_duration // 60)
            total_seconds = int(self.video_duration % 60)
            painter.drawText(width - 30, height - 5, f"{total_minutes:02d}:{total_seconds:02d}")
    
    def mousePressEvent(self, event):
        """鼠标点击事件 - 跳转到对应时间"""
        if event.button() == Qt.LeftButton and self.video_duration > 0:
            width = self.width() - 20
            x_offset = 10
            click_x = event.x() - x_offset
            
            if 0 <= click_x <= width:
                # 计算点击位置对应的时间
                click_time = (click_x / width) * self.video_duration
                target_frame = int(click_time * self.fps)
                
                # 查找父级分析器并调用跳转方法
                parent = self.parent()
                while parent and not hasattr(parent, 'seek_to_frame'):
                    parent = parent.parent()
                if parent and hasattr(parent, 'seek_to_frame'):
                    parent.seek_to_frame(target_frame)
    
    def mouseMoveEvent(self, event):
        """鼠标移动事件 - 显示悬停提示"""
        if self.video_duration > 0:
            width = self.width() - 20
            x_offset = 10
            mouse_x = event.x() - x_offset
            
            if 0 <= mouse_x <= width:
                # 检查是否悬停在事件标记上
                hover_time = (mouse_x / width) * self.video_duration
                
                for event_data in self.events:
                    time_diff = abs(event_data['time'] - hover_time)
                    if time_diff <= 1.0:  # 1秒容差
                        # 显示事件信息
                        tooltip_text = f"时间: {int(event_data['time'] // 60):02d}:{int(event_data['time'] % 60):02d}\n"
                        tooltip_text += f"球员: {event_data['player_id']}号\n"
                        tooltip_text += f"动作: {event_data['action']}\n"
                        tooltip_text += f"帧号: {event_data['frame']}"
                        
                        QToolTip.showText(event.globalPos(), tooltip_text)
                        return
                
                # 如果没有悬停在事件上，显示时间信息
                time_str = f"时间: {int(hover_time // 60):02d}:{int(hover_time % 60):02d}"
                QToolTip.showText(event.globalPos(), time_str)

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = VolleyballAnalyzer()
    window.show()
    sys.exit(app.exec_())